# mail crawler
import tkinter
from tkinter import filedialog
import tkinter as tk
import openpyxl as op
import re
import sys
##import urllib3
##import certifi
##import io
##import codecs
##import json
##import ssl
##import os
##from OpenSSL import SSL

from selenium import webdriver
import time

#import urllib.request as urllib2
global loc, lis, sh


##def parseAddress(link):
##
##    try:
##        #http = urllib3.PoolManager(cert_reqs='CERT_REQUIRED',ca_certs='cacert.pem')
##        http = urllib3.PoolManager(cert_reqs='CERT_REQUIRED',ca_certs=certifi.where())
##        #r = (http.request('GET', link,preload_content=False,decode_content ='cp437'))
##        r = (http.request('GET', link,preload_content=False,decode_content ='utf-8'))
##        
##        r.auto_close = False
##        
##        #print("raju read"+str(r.read()))
##        g=str(r.read())
##        add = re.findall( '''[a-z0-9!#$%&'*+/=?^_`{|}~-]+(?:\.[a-z0-9!#$%&'*+/=?^_`{|}~-]+)*@(?:[a-z0-9](?:[a-z0-9-]*[a-z0-9])?\.)+[a-z0-9](?:[a-z0-9-]*[a-z0-9])?''', g, flags=re.IGNORECASE)
##        #print(add)
##        if(str(add)!='[]'):
##            #print(add[0])
##            return str(add[0])
##
##            
##    except IOError as w:
##        pass
##        #print("IOerr => "+str(w))
##    
##
##    except Exception as e:
##        print('Unsuccesful because => '+str(e))
##        print(type(e).__name__,__file__,e.__traceback__.tb_lineno)
##        return ''
##
##    

def web_thing(link):
    bha=''
    
    #driver =webdriver.Chrome('C:/Project OverLoad/Automation/chromedriver.exe')
    driver =webdriver.Chrome('chromedriver.exe')
    try:
        driver.get(link)
        time.sleep(1)
##        print(str(page_is_loading(driver)))
##        while not page_is_loading(driver):
##            print('blah '+str(page_is_loading(driver)))
##            continue

        html=str(driver.page_source)
        #print(str(html))
        add = re.findall( '''[a-z0-9!#$%&'*+/=?^_`{|}~-]+(?:\.[a-z0-9!#$%&'*+/=?^_`{|}~-]+)*@(?:[a-z0-9](?:[a-z0-9-]*[a-z0-9])?\.)+[a-z0-9](?:[a-z0-9-]*[a-z0-9])?''', html, flags=re.IGNORECASE)
    #print(add)
    except Exception as e:
        
        print(str(e))
        print(type(e).__name__,__file__,e.__traceback__.tb_lineno)
        bha=''
        #return ''
    if (len(driver.window_handles)>1):
            window_name = driver.window_handles[0]
            driver.switch_to.window(window_name=window_name)
            driver.close()
            
    try:
        if(str(add)!='[]'):
            #print(add[0])
            bha=str(add[0])
            #return str(add[0])
        
        else:
            bha=''
            #return ''

    except Exception as e:
        bha=''
        #return ''
        
    try:
        driver.quit()
    except Exception:
        pass
    return bha
    

def inside_data(k):
    global loc, lis, sh
    for i in range(1, k):
        #print(sh['A'+str(i)].value)
        if(sh['A'+str(i)].value == None or (sh['A'+str(i)].value).strip() == ""):
            break
        print(str(i)+'  '+(sh['A'+str(i)].value).strip())
        #g = parseAddress((sh['A'+str(i)].value).strip())
        g = web_thing((sh['A'+str(i)].value).strip())
        if(g=='' or g==None):
            dab='B'+str(i)
            sh[dab]=''
            print('NO mailid found')
            print()
        else:
            if '/' in g or '\ ' in g:
                kk=sh.cell(row=i ,column=2)
                kk.value=str('')
                #dab='B'+str(i)
                #sh[dab]=str('')
                lis.save(loc)
                print('invalid one found')
                
            else:
                #dab='B'+str(i)
                #sh[dab]=str(g)
                kk=sh.cell(row=i ,column=2)
                kk.value=str(g)
                print('mail => '+str(g))
                print()
                lis.save(loc)
                
                


def read():
    global loc, lis, sh
    root = tk.Tk()
    loc = filedialog.askopenfilename(
        initialdir="/", title="Select the fle", filetypes=(("XLSX", "*.xlsx*"), ("all files", "*.*")))
    root.destroy()
    lis = op.load_workbook(loc)
    sh = lis['Sheet1']
    print(sh.max_row)
    inside_data(sh.max_row)


read()
